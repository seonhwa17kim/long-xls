"""Export helpers: xlsx (default), csv, parquet, and JSON schema."""

from __future__ import annotations

import csv
import io
import json
import sys
from pathlib import Path
from typing import Any, Callable

from long_xls.parser import SheetData, _OPCODE_NAME

ProgressFn = Callable[[int, int], None]


def _null_progress(cur: int, total: int) -> None:
    pass


# -- Schema ----------------------------------------------------------------

def schema_json(sheet: SheetData) -> dict[str, Any]:
    """Return a JSON-serialisable dict describing the file structure."""
    col_types: list[dict[str, Any]] = []
    for c in range(sheet.num_columns):
        vals = sheet.columns.get(c, [])
        sample = next((v for v in vals if v is not None), None)
        if isinstance(sample, str):
            dtype = "string"
        elif isinstance(sample, int):
            dtype = "integer"
        elif isinstance(sample, float):
            dtype = "float"
        else:
            dtype = "unknown"
        col_types.append({
            "index": c,
            "name": sheet.headers[c],
            "type": dtype,
            "non_null_count": sum(1 for v in vals if v is not None),
        })

    return {
        "file": sheet.path,
        "file_size": sheet.file_size,
        "encoding": sheet.encoding,
        "num_columns": sheet.num_columns,
        "num_data_rows": sheet.num_data_rows,
        "row_limit_exceeded": any(w > 0 for w in sheet.wraps_per_col.values()),
        "wraps": dict(sheet.wraps_per_col),
        "columns": col_types,
        "record_types": {
            _OPCODE_NAME.get(op, f"0x{op:04X}"): cnt
            for op, cnt in sorted(sheet.record_counts.items())
        },
        "warnings": sheet.warnings,
    }


def write_schema(sheet: SheetData, dest: str | Path) -> Path:
    """Write schema JSON to a file."""
    dest = Path(dest)
    dest.write_text(
        json.dumps(schema_json(sheet), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return dest


# -- Row iterator (shared by all exporters) --------------------------------

def _iter_rows(sheet: SheetData):
    """Yield one row at a time as a list of values."""
    ncols = sheet.num_columns
    cols = [sheet.columns.get(c, []) for c in range(ncols)]
    for i in range(sheet.num_data_rows):
        yield [cols[c][i] if i < len(cols[c]) else None for c in range(ncols)]


# -- Exporters with progress -----------------------------------------------

def to_xlsx(
    sheet: SheetData,
    dest: str | Path,
    progress: ProgressFn = _null_progress,
) -> Path:
    """Write to XLSX using openpyxl write_only mode for speed + progress."""
    from openpyxl import Workbook

    dest = Path(dest)
    total = sheet.num_data_rows

    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    # Header
    ws.append(sheet.headers)

    # Data rows
    interval = max(total // 50, 1)
    for i, row in enumerate(_iter_rows(sheet)):
        ws.append(row)
        if i % interval == 0:
            progress(i, total)

    progress(total, total)
    wb.save(str(dest))
    wb.close()
    return dest


def to_csv(
    sheet: SheetData,
    dest: str | Path,
    progress: ProgressFn = _null_progress,
) -> Path:
    """Write to CSV with UTF-8 BOM and progress."""
    dest = Path(dest)
    total = sheet.num_data_rows
    interval = max(total // 50, 1)

    with open(dest, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(sheet.headers)
        for i, row in enumerate(_iter_rows(sheet)):
            writer.writerow(row)
            if i % interval == 0:
                progress(i, total)

    progress(total, total)
    return dest


def to_parquet(
    sheet: SheetData,
    dest: str | Path,
    progress: ProgressFn = _null_progress,
) -> Path:
    """Write to Parquet. Progress is coarse (build DF -> write)."""
    import pandas as pd

    dest = Path(dest)
    total = 3  # 3 steps: build dict, create df, write

    data = {
        sheet.headers[c]: sheet.columns.get(c, [None] * sheet.num_data_rows)
        for c in range(sheet.num_columns)
    }
    progress(1, total)

    df = pd.DataFrame(data)
    progress(2, total)

    df.to_parquet(str(dest), index=False, engine="pyarrow")
    progress(total, total)
    return dest


FORMATS = {
    "xlsx": to_xlsx,
    "csv": to_csv,
    "parquet": to_parquet,
}
